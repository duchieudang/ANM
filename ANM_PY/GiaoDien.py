import tkinter as tk
from tkinter import messagebox, ttk, filedialog
from docx import Document
import fitz  # PyMuPDF
from openpyxl import load_workbook
import random
from dsa import is_prime, calculate_g, generate_k, sign, verify


class DSAGui:
    def __init__(self, root):
        self.root = root
        root.title("DSA Digital Signature Tool")
        root.geometry("1200x750")

        self.setup_style()

        self.r = None
        self.s = None
        self.message = None

        self.p_var = tk.StringVar()
        self.q_var = tk.StringVar()
        self.h_var = tk.StringVar()
        self.g_var = tk.StringVar()
        self.x_var = tk.StringVar()
        self.y_var = tk.StringVar()
        self.k_var = tk.StringVar()
        self.hash_alg_var = tk.StringVar(value="SHA-256")  # ✅

        self.setup_widgets()

    def setup_style(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TLabel", font=('Segoe UI', 10), foreground="#333")
        style.configure("TEntry", font=('Segoe UI', 10), fieldbackground="#fff")
        style.configure("TFrame", background="#f5f5dc")
        style.configure("TLabelframe.Label", font=('Segoe UI', 11, 'bold'), foreground="#333")

        style.configure("Param.TLabelframe", background="#e6e6fa")
        style.configure("Sign.TLabelframe", background="#d4edda")
        style.configure("Verify.TLabelframe", background="#ffe5d9")

        style.configure("Generate.TButton", font=('Segoe UI', 10), background="#e2e3f5", foreground="#333", padding=6)
        style.map("Generate.TButton", background=[("active", "#d1d2f0")])
        style.configure("Reset.TButton", font=('Segoe UI', 10), background="#f8d7da", foreground="#333", padding=6)
        style.map("Reset.TButton", background=[("active", "#f5c6cb")])
        style.configure("Sign.TButton", font=('Segoe UI', 10), background="#c3e6cb", foreground="#333", padding=6)
        style.map("Sign.TButton", background=[("active", "#b1d8b7")])
        style.configure("Verify.TButton", font=('Segoe UI', 10), background="#ffddd2", foreground="#333", padding=6)
        style.map("Verify.TButton", background=[("active", "#ffccb3")])

    def setup_widgets(self):
        main_frame = ttk.Frame(self.root)
        main_frame.pack(expand=True, fill='both', padx=5, pady=5)

        for i in range(3):
            main_frame.grid_columnconfigure(i, weight=1, uniform="col")
            main_frame.grid_rowconfigure(0, weight=1)

        self.create_param_section(main_frame, 0)
        self.create_sign_section(main_frame, 1)
        self.create_verify_section(main_frame, 2)

    def create_param_section(self, parent, column):
        frame = ttk.LabelFrame(parent, text="Tham số DSA", padding=10, style="Param.TLabelframe")
        frame.grid(row=0, column=column, sticky="nsew", padx=5, pady=5)

        fields = [
            ("p:", self.p_var),
            ("q:", self.q_var),
            ("h:", self.h_var),
            ("g:", self.g_var),
            ("x (private key):", self.x_var),
            ("y (public key):", self.y_var),
            ("k (tùy chọn):", self.k_var),
        ]

        for i, (label, var) in enumerate(fields):
            ttk.Label(frame, text=label).grid(row=i, column=0, sticky="e", padx=5, pady=2)
            ttk.Entry(frame, textvariable=var, width=20).grid(row=i, column=1, sticky="w", padx=5, pady=2)

        ttk.Label(frame, text="Hàm băm:").grid(row=len(fields), column=0, sticky="e", padx=5, pady=2)
        ttk.Combobox(frame, textvariable=self.hash_alg_var,
                     values=["SHA-1", "SHA-256", "SHA-512"],
                     state="readonly", width=17).grid(row=len(fields), column=1, sticky="w", padx=5, pady=2)

        ttk.Button(frame, text="🎲 Sinh ngẫu nhiên", command=self.generate_all_params, style="Generate.TButton")\
            .grid(row=len(fields)+1, column=0, columnspan=2, pady=5, sticky="ew")
        ttk.Button(frame, text="🔄 Reset", command=self.reset_params, style="Reset.TButton")\
            .grid(row=len(fields)+2, column=0, columnspan=2, pady=5, sticky="ew")

    def create_sign_section(self, parent, column):
        frame = ttk.LabelFrame(parent, text="Ký nội dung", padding=10, style="Sign.TLabelframe")
        frame.grid(row=0, column=column, sticky="nsew", padx=5, pady=5)

        ttk.Button(frame, text="📂 Mở file (Word, PDF, Excel)", command=lambda: self.load_file(self.text_input)).pack(pady=5, fill='x')

        ttk.Label(frame, text="Nội dung cần ký:").pack(anchor='w')
        self.text_input = tk.Text(frame, height=10, wrap='word')
        self.text_input.pack(fill="both", expand=True, pady=5)

        ttk.Button(frame, text="✍️ Tạo chữ ký", command=self.sign_message, style="Sign.TButton").pack(pady=5, fill='x')
        ttk.Button(frame, text="➡️ Chuyển sang xác thực", command=self.transfer_to_verify, style="Verify.TButton")\
            .pack(pady=5, fill='x')

        ttk.Label(frame, text="Kết quả (r, s):").pack(anchor='w')
        self.signature_text = tk.Text(frame, height=5, wrap='word', state='disabled')
        self.signature_text.pack(fill="both", expand=True, pady=5)

    def create_verify_section(self, parent, column):
        frame = ttk.LabelFrame(parent, text="Xác thực chữ ký", padding=10, style="Verify.TLabelframe")
        frame.grid(row=0, column=column, sticky="nsew", padx=5, pady=5)

        ttk.Button(frame, text="📂 Mở file (Word, PDF, Excel)", command=lambda: self.load_file(self.verify_text_input)).pack(pady=5, fill='x')

        ttk.Label(frame, text="Nội dung cần xác thực:").pack(anchor='w')
        self.verify_text_input = tk.Text(frame, height=10, wrap='word')
        self.verify_text_input.pack(fill="both", expand=True, pady=5)

        ttk.Label(frame, text="Nhập r:").pack(anchor='w')
        self.r_entry = ttk.Entry(frame)
        self.r_entry.pack(fill='x', pady=2)

        ttk.Label(frame, text="Nhập s:").pack(anchor='w')
        self.s_entry = ttk.Entry(frame)
        self.s_entry.pack(fill='x', pady=2)

        ttk.Button(frame, text="✅ Xác thực", command=self.verify_signature, style="Verify.TButton")\
            .pack(pady=10, fill='x')
        self.result_label = tk.Label(frame, text="", font=('Segoe UI', 11, 'bold'), fg="#333")
        self.result_label.pack(fill='x')

    def load_file(self, text_widget):
        file_path = filedialog.askopenfilename(
            title="Chọn file (Word, PDF, Excel)",
            filetypes=[
                ("All files", "*.docx *.pdf *.xlsx"),
                ("Word Documents", "*.docx"),
                ("PDF Files", "*.pdf"),
                ("Excel Files", "*.xlsx"),
            ]
        )
        if not file_path:
            return
        try:
            if file_path.endswith(".docx"):
                doc = Document(file_path)
                full_text = "\n".join([para.text for para in doc.paragraphs])
            elif file_path.endswith(".pdf"):
                doc = fitz.open(file_path)
                full_text = "\n".join([page.get_text() for page in doc])
            elif file_path.endswith(".xlsx"):
                wb = load_workbook(filename=file_path)
                sheet = wb.active
                full_text = ""
                for row in sheet.iter_rows(values_only=True):
                    full_text += "\t".join([str(cell) if cell is not None else "" for cell in row]) + "\n"
            else:
                messagebox.showwarning("Định dạng file", "Chỉ hỗ trợ .docx, .pdf, .xlsx")
                return

            text_widget.delete("1.0", tk.END)
            text_widget.insert(tk.END, full_text)

        except Exception as e:
            messagebox.showerror("Lỗi mở file", str(e))

    def generate_all_params(self):
        self.q = random.choice([101, 103, 107])
        self.p = self.q * random.randint(70, 100) + 1
        while not is_prime(self.p):
            self.p += self.q
        self.h = random.randint(2, self.p - 2)
        self.g = calculate_g(self.p, self.q, self.h)
        self.x = random.randint(1, self.q - 1)
        self.y = pow(self.g, self.x, self.p)
        self.k = generate_k(self.q)

        self.p_var.set(str(self.p))
        self.q_var.set(str(self.q))
        self.h_var.set(str(self.h))
        self.g_var.set(str(self.g))
        self.x_var.set(str(self.x))
        self.y_var.set(str(self.y))
        self.k_var.set(str(self.k))

    def reset_params(self):
        for var in [self.p_var, self.q_var, self.h_var, self.g_var, self.x_var, self.y_var, self.k_var]:
            var.set("")
        self.hash_alg_var.set("SHA-256")
        self.r = None
        self.s = None
        self.message = None
        self.signature_text.config(state='normal')
        self.signature_text.delete("1.0", tk.END)
        self.signature_text.config(state='disabled')
        self.result_label.config(text="")
        self.text_input.delete("1.0", tk.END)
        self.verify_text_input.delete("1.0", tk.END)
        self.r_entry.delete(0, tk.END)
        self.s_entry.delete(0, tk.END)

    def transfer_to_verify(self):
        if self.r is None or self.s is None or self.message is None:
            messagebox.showwarning("Cảnh báo", "Hãy ký nội dung trước!")
            return

        self.verify_text_input.delete("1.0", tk.END)
        self.verify_text_input.insert(tk.END, self.message)
        self.r_entry.delete(0, tk.END)
        self.r_entry.insert(0, str(self.r))
        self.s_entry.delete(0, tk.END)
        self.s_entry.insert(0, str(self.s))

        messagebox.showinfo("Đã chuyển", "Nội dung và chữ ký đã được chuyển sang phần xác thực.")

    def sign_message(self):
        try:
            message = self.text_input.get("1.0", tk.END).strip()
            if not message:
                messagebox.showwarning("Cảnh báo", "Vui lòng nhập nội dung cần ký.")
                return

            self.p = int(self.p_var.get())
            self.q = int(self.q_var.get())
            self.h = int(self.h_var.get())
            self.g = int(self.g_var.get())
            self.x = int(self.x_var.get())
            self.y = int(self.y_var.get())

            if not is_prime(self.p) or not is_prime(self.q):
                raise ValueError("p và q phải là số nguyên tố!")

            k_input = self.k_var.get().strip()
            if k_input:
                self.k = int(k_input)
                if not (0 < self.k < self.q):
                    raise ValueError("k phải nằm trong khoảng (0, q)!")
            else:
                self.k = generate_k(self.q)

            hash_alg = self.hash_alg_var.get()
            self.r, self.s = sign(message, self.p, self.q, self.g, self.x, self.k, hash_alg)
            self.message = message

            self.signature_text.config(state='normal')
            self.signature_text.delete("1.0", tk.END)
            self.signature_text.insert(tk.END, f"r = {self.r}\ns = {self.s}")
            self.signature_text.config(state='disabled')
        except Exception as e:
            messagebox.showerror("Lỗi khi ký", str(e))

    def verify_signature(self):
        try:
            message = self.verify_text_input.get("1.0", tk.END).strip()
            r2 = int(self.r_entry.get().strip())
            s2 = int(self.s_entry.get().strip())

            self.p = int(self.p_var.get())
            self.q = int(self.q_var.get())
            self.h = int(self.h_var.get())
            self.g = int(self.g_var.get())
            self.y = int(self.y_var.get())

            errors = []
            if self.r != r2:
                errors.append("⚠️ r đã bị thay đổi!")
            if self.s != s2:
                errors.append("⚠️ s đã bị thay đổi!")
            if self.message != message:
                errors.append("⚠️ Văn bản đã bị thay đổi!")

            hash_alg = self.hash_alg_var.get()
            is_valid = verify(message, self.p, self.q, self.g, self.y, r2, s2, hash_alg)

            if is_valid:
                self.result_label.config(text="✅ Chữ ký hợp lệ", fg="green")
            else:
                self.result_label.config(text="❌ Chữ ký không hợp lệ", fg="red")

            if errors:
                messagebox.showwarning("Cảnh báo", "\n".join(errors))
        except Exception as e:
            messagebox.showerror("Lỗi khi xác thực", str(e))


if __name__ == "__main__":
    root = tk.Tk()
    app = DSAGui(root)
    root.mainloop()
