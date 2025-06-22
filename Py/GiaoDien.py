import tkinter as tk
from tkinter import messagebox, ttk, filedialog
from docx import Document
import fitz  # PyMuPDF
from openpyxl import load_workbook
import random
from dsa import is_prime, calculate_g, generate_k, sign, verify
from PIL import Image, ImageTk
import os  # Thêm để debug đường dẫn
import random
from sympy import isprime
class DSAGui:
    def __init__(self, root):
        self.root = root
        root.title("DSA Digital Signature Tool")
        root.geometry("1200x750")

        self.setup_style()
        self.load_icons()  # Nạp icon

        self.r = None
        self.s = None
        self.message = None

        self.p_var = tk.StringVar()
        self.q_var = tk.StringVar()
        self.h_var = tk.StringVar()
        self.k_var = tk.StringVar()
        self.x_var = tk.StringVar()
        self.g_var = tk.StringVar()
        self.y_var = tk.StringVar()

        self.setup_widgets()

    def setup_style(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TLabel", font=('Segoe UI', 10), foreground="black")
        style.configure("TEntry", font=('Segoe UI', 10), fieldbackground="white")
        style.configure("TFrame", background="#f5f5dc")
        style.configure("TLabelframe.Label", font=('Segoe UI', 11, 'bold'), foreground="black")

        style.configure("Param.TLabelframe", background="#e6e6fa")
        style.configure("Sign.TLabelframe", background="#d4edda")
        style.configure("Verify.TLabelframe", background="#f5f5dc")

        # Custom styles for all buttons with lighter colors
        style.configure("Generate.TButton", font=('Segoe UI', 11), background="#66BB6A", foreground="black", padding=5)
        style.map("Generate.TButton", background=[("active", "#43A047")])
        style.configure("Reset.TButton", font=('Segoe UI', 11), background="#EF9A9A", foreground="black", padding=5)
        style.map("Reset.TButton", background=[("active", "#AD1457")])
        style.configure("Check.TButton", font=('Segoe UI', 11), background="#42A5F5", foreground="black", padding=5)
        style.map("Check.TButton", background=[("active", "#1976D2")])
        style.configure("SignOpen.TButton", font=('Segoe UI', 11), background="#66BB6A", foreground="black", padding=5)
        style.map("SignOpen.TButton", background=[("active", "#43A047")])
        style.configure("SignCreate.TButton", font=('Segoe UI', 11), background="#EF9A9A", foreground="black", padding=5)
        style.map("SignCreate.TButton", background=[("active", "#AD1457")])
        style.configure("SignTransfer.TButton", font=('Segoe UI', 11), background="#fccc7e", foreground="black", padding=5)
        style.map("SignTransfer.TButton", background=[("active", "#F57C00")])
        style.configure("VerifyOpen.TButton", font=('Segoe UI', 11), background="#42A5F5", foreground="black", padding=5)
        style.map("VerifyOpen.TButton", background=[("active", "#1976D2")])
        style.configure("Verify.TButton", font=('Segoe UI', 11), background="#fccc7e", foreground="black", padding=5)
        style.map("Verify.TButton", background=[("active", "#F57C00")])
        style.configure("Save.TButton", font=('Segoe UI', 11), background="#66BB6A", foreground="black", padding=5)
        style.map("Save.TButton", background=[("active", "#43A047")])
        style.configure("Import.TButton", font=('Segoe UI', 11), background="#EF9A9A", foreground="black", padding=5)
        style.map("Import.TButton", background=[("active", "#AD1457")])

        # Custom style for readonly entries (g and y) with darker gray background
        style.configure("Readonly.TEntry", fieldbackground="#A9A9A9", foreground="black")

    def load_icons(self):
        current_dir = os.path.dirname(__file__)
        image_dir = os.path.join(current_dir, "Anh")

        self.reset = ImageTk.PhotoImage(Image.open(os.path.join(image_dir, "reset.png")).resize((25, 25), Image.LANCZOS))
        self.dice = ImageTk.PhotoImage(Image.open(os.path.join(image_dir, "dice.png")).resize((25, 25), Image.LANCZOS))
        self.check = ImageTk.PhotoImage(Image.open(os.path.join(image_dir, "check1.png")).resize((25, 25), Image.LANCZOS))
        self.book = ImageTk.PhotoImage(Image.open(os.path.join(image_dir, "book.png")).resize((30, 30), Image.LANCZOS))
        self.arrow = ImageTk.PhotoImage(Image.open(os.path.join(image_dir, "arrow.png")).resize((30, 30), Image.LANCZOS))
        self.edit = ImageTk.PhotoImage(Image.open(os.path.join(image_dir, "check.png")).resize((30, 30), Image.LANCZOS))
        self.save = ImageTk.PhotoImage(Image.open(os.path.join(image_dir, "save.png")).resize((30, 30), Image.LANCZOS))
        self.import_icon = ImageTk.PhotoImage(Image.open(os.path.join(image_dir, "import.png")).resize((30, 30), Image.LANCZOS))

    def setup_widgets(self):
        main_frame = ttk.Frame(self.root)
        main_frame.pack(expand=True, fill='both', padx=5, pady=5)

        for i in range(3):
            main_frame.grid_columnconfigure(i, weight=1, uniform="col")
            main_frame.grid_rowconfigure(0, weight=1)

        self.create_param_section(main_frame, 0)
        self.create_sign_section(main_frame, 1)
        self.create_verify_section(main_frame, 2)

    def create_param_section(self, parent, column):
        frame = ttk.LabelFrame(parent, text="Tham số DSA", padding=12, style="Param.TLabelframe")
        frame.grid(row=0, column=column, sticky="nsew", padx=5, pady=5)

        fields = [
            ("số nguyên tố p:", self.p_var),
            ("số nguyên tố q:", self.q_var),
            ("h (1<h<p-1):", self.h_var),
            ("k (0<k<q):", self.k_var),
            ("private key x(0<x<q):", self.x_var),
        ]

        for i, (label, var) in enumerate(fields):
            ttk.Label(frame, text=label).grid(row=i, column=0, sticky="e", padx=5, pady=2)
            ttk.Entry(frame, textvariable=var, width=20).grid(row=i, column=1, sticky="w", padx=5, pady=2)

        # Hiển thị g và y (readonly)
        ttk.Label(frame, text="g=h^((p-1)/q) mod p:").grid(row=len(fields), column=0, sticky="e", padx=5, pady=2)
        ttk.Entry(frame, textvariable=self.g_var, width=20, state="readonly", style="Readonly.TEntry")\
            .grid(row=len(fields), column=1, sticky="w", padx=5, pady=2)
        ttk.Label(frame, text="public key y :").grid(row=len(fields)+1, column=0, sticky="e", padx=5, pady=2)
        ttk.Entry(frame, textvariable=self.y_var, width=20, state="readonly", style="Readonly.TEntry")\
            .grid(row=len(fields)+1, column=1, sticky="w", padx=5, pady=2)

        # Các nút chức năng
        ttk.Button(frame, text="Sinh ngẫu nhiên", image=self.dice, compound="left",
                   command=self.generate_all_params, style="Generate.TButton")\
            .grid(row=len(fields)+2, column=0, columnspan=2, pady=5, sticky="ew")
        ttk.Button(frame, text="Reset", image=self.reset, compound="left",
                   command=self.reset_params, style="Reset.TButton")\
            .grid(row=len(fields)+3, column=0, columnspan=2, pady=5, sticky="ew")
        ttk.Button(frame, text="Kiểm tra", image=self.check, compound="left",
                   command=self.check_params, style="Check.TButton")\
            .grid(row=len(fields)+4, column=0, columnspan=2, pady=5, sticky="ew")

    def create_sign_section(self, parent, column):
        frame = ttk.LabelFrame(parent, text="Ký nội dung", padding=12, style="Sign.TLabelframe")
        frame.grid(row=0, column=column, sticky="nsew", padx=5, pady=5)

        ttk.Button(frame, text="Mở file (Word, PDF, Excel)", image=self.book, compound="left",
                   command=lambda: self.load_file(self.text_input), style="SignOpen.TButton").pack(pady=5, fill='x')

        ttk.Label(frame, text="Nội dung cần ký:").pack(anchor='w')
        self.text_input = tk.Text(frame, height=10, wrap='word')
        self.text_input.pack(fill="both", expand=True, pady=5)

        # Thêm lựa chọn hàm băm
        ttk.Label(frame, text="Thuật toán băm:").pack(anchor='w', pady=2)
        self.hash_alg_var = tk.StringVar(value="SHA-256")
        hash_combo = ttk.Combobox(frame, textvariable=self.hash_alg_var,
                                  values=["SHA-1", "SHA-256", "SHA-512"], state="readonly", width=17)
        hash_combo.pack(fill='x', pady=2)

        ttk.Button(frame, text="Tạo chữ ký", image=self.edit, compound="left", command=self.sign_message, style="SignCreate.TButton").pack(pady=5, fill='x')
        ttk.Button(frame, text="Lưu chữ ký", image=self.save, compound="left", command=self.save_signature, style="Save.TButton").pack(pady=5, fill='x')
        ttk.Button(frame, text="Chuyển sang xác thực", image=self.arrow, compound="left", command=self.transfer_to_verify, style="SignTransfer.TButton").pack(pady=5, fill='x')

        ttk.Label(frame, text="Kết quả (r, s):").pack(anchor='w')
        self.signature_text = tk.Text(frame, height=5, wrap='word', state='disabled')
        self.signature_text.pack(fill="both", expand=True, pady=5)

    def create_verify_section(self, parent, column):
        frame = ttk.LabelFrame(parent, text="Xác thực chữ ký", padding=12, style="Verify.TLabelframe")
        frame.grid(row=0, column=column, sticky="nsew", padx=5, pady=5)

        ttk.Button(frame, text="Mở file (Word, PDF, Excel)", image=self.book, compound="left", command=lambda: self.load_file(self.verify_text_input), style="VerifyOpen.TButton").pack(pady=5, fill='x')

        ttk.Label(frame, text="Nội dung cần xác thực:").pack(anchor='w')
        self.verify_text_input = tk.Text(frame, height=10, wrap='word')
        self.verify_text_input.pack(fill="both", expand=True, pady=5)

        ttk.Label(frame, text="Nhập r:").pack(anchor='w')
        self.r_entry = ttk.Entry(frame)
        self.r_entry.pack(fill='x', pady=2)

        ttk.Label(frame, text="Nhập s:").pack(anchor='w')
        self.s_entry = ttk.Entry(frame)
        self.s_entry.pack(fill='x', pady=2)

        ttk.Button(frame, text="Nhập chữ ký", image=self.import_icon, compound="left", command=self.import_signature, style="Import.TButton").pack(pady=5, fill='x')
        ttk.Button(frame, text="Xác thực", image=self.check, compound="left", command=self.verify_signature, style="Verify.TButton")\
            .pack(pady=10, fill='x')
        self.result_label = tk.Label(frame, text="", font=('Segoe UI', 11, 'bold'), fg="#333")
        self.result_label.pack(fill='x')

    def load_file(self, text_widget):
        file_path = filedialog.askopenfilename(
            title="Chọn file (Word, PDF, Excel)",
            filetypes=[
                ("All files", "*.docx *.pdf *.xlsx"),
                ("Word Documents", "*.docx"),
                ("PDF Files", "*.pdf"),
                ("Excel Files", "*.xlsx"),
            ]
        )
        if not file_path:
            return
        try:
            if file_path.endswith(".docx"):
                doc = Document(file_path)
                full_text = "\n".join([para.text for para in doc.paragraphs])
                text_widget.delete("1.0", tk.END)
                text_widget.insert(tk.END, full_text)
            elif file_path.endswith(".pdf"):
                doc = fitz.open(file_path)
                content = "\n".join([page.get_text() for page in doc])
                text_widget.delete("1.0", tk.END)
                text_widget.insert(tk.END, content)
            elif file_path.endswith(".xlsx"):
                wb = load_workbook(filename=file_path)
                sheet = wb.active
                content = ""
                for row in sheet.iter_rows(values_only=True):
                    content += "\t".join([str(cell) if cell is not None else "" for cell in row]) + "\n"
                text_widget.delete("1.0", tk.END)
                text_widget.insert(tk.END, content)
            else:
                messagebox.showwarning("Định dạng file", "Chỉ hỗ trợ file .docx, .pdf, .xlsx")
        except Exception as e:
            messagebox.showerror("Lỗi mở file", str(e))
    def calculate_g(p, q, h):
        return pow(h, (p - 1) // q, p)

    def generate_k(q):
        return random.randint(1, q - 1)

    def generate_large_prime(self, bits):
        while True:
            n = random.getrandbits(bits)
            n |= 1 
            if isprime(n):
                return n

    def generate_all_params(self):
        self.q = self.generate_large_prime(128)  # thể hiện số bit của nó ở đây

        k = random.randint(2**5, 2**8) 
        self.p = self.q * k + 1
        while not isprime(self.p):
            self.p += self.q

        self.h = random.randint(2, self.p - 2)
        self.g = calculate_g(self.p, self.q, self.h)
        self.x = random.randint(1, self.q - 1)
        self.k = generate_k(self.q)
        self.y = pow(self.g, self.x, self.p)

        self.p_var.set(str(self.p))
        self.q_var.set(str(self.q))
        self.h_var.set(str(self.h))
        self.k_var.set(str(self.k))
        self.x_var.set(str(self.x))
        self.g_var.set(str(self.g))
        self.y_var.set(str(self.y))


    def reset_params(self):
        for var in [self.p_var, self.q_var, self.h_var, self.k_var, self.x_var, self.g_var, self.y_var]:
            var.set("")
        self.r = None
        self.s = None
        self.message = None
        self.signature_text.config(state='normal')
        self.signature_text.delete("1.0", tk.END)
        self.signature_text.config(state='disabled')
        self.result_label.config(text="")
        self.text_input.delete("1.0", tk.END)
        self.verify_text_input.delete("1.0", tk.END)
        self.r_entry.delete(0, tk.END)
        self.s_entry.delete(0, tk.END)

    def check_params(self):
        try:
            p = int(self.p_var.get()) if self.p_var.get() else 0
            q = int(self.q_var.get()) if self.q_var.get() else 0
            h = int(self.h_var.get()) if self.h_var.get() else 0
            k = int(self.k_var.get()) if self.k_var.get() else 0
            x = int(self.x_var.get()) if self.x_var.get() else 0

            # Kiểm tra q là số nguyên tố
            if not is_prime(q):
                messagebox.showerror("Lỗi", "q không phải là số nguyên tố!")
                return

            # Kiểm tra p là số nguyên tố và (p - 1) chia hết cho q
            if not is_prime(p):
                messagebox.showerror("Lỗi", "p không phải là số nguyên tố!")
                return
            if (p - 1) % q != 0:
                messagebox.showerror("Lỗi", "p phải thỏa mãn (p - 1) chia hết cho q!")
                return

            # Kiểm tra h nằm trong khoảng (1, p - 1)
            if h <= 1 or h >= p - 1:
                messagebox.showerror("Lỗi", "h phải là số nguyên nằm trong khoảng (1, p - 1)!")
                return

            # Tính g và kiểm tra g > 1
            g = calculate_g(p, q, h)
            if g <= 1:
                messagebox.showerror("Lỗi", "g phải lớn hơn 1!")
                return
            self.g_var.set(str(g))

            # Kiểm tra x nằm trong khoảng (0, q)
            if x <= 0 or x >= q:
                messagebox.showerror("Lỗi", "x (khóa bí mật) phải nằm trong khoảng (0, q)!")
                return

            # Kiểm tra k nằm trong khoảng (0, q)
            if k <= 0 or k >= q:
                messagebox.showerror("Lỗi", "k (số ngẫu nhiên bí mật) phải nằm trong khoảng (0, q)!")
                return

            # Tính y và gán giá trị
            y = pow(g, x, p)
            self.y_var.set(str(y))

            messagebox.showinfo("Thành công", "Kiểm tra thành công! Các tham số g và y đã được tính toán.")
        except ValueError:
            messagebox.showerror("Lỗi", "Vui lòng nhập đầy đủ và đúng định dạng số nguyên!")
        except Exception as e:
            messagebox.showerror("Lỗi", str(e))

    def transfer_to_verify(self):
        if self.r is None or self.s is None or self.message is None:
            messagebox.showwarning("Cảnh báo", "Hãy ký nội dung trước!")
            return

        self.verify_text_input.delete("1.0", tk.END)
        self.verify_text_input.insert(tk.END, self.message)
        self.r_entry.delete(0, tk.END)
        self.r_entry.insert(0, str(self.r))
        self.s_entry.delete(0, tk.END)
        self.s_entry.insert(0, str(self.s))

        messagebox.showinfo("Đã chuyển", "Nội dung và chữ ký đã được chuyển sang phần xác thực.")

    def sign_message(self):
        try:
            message = self.text_input.get("1.0", tk.END).strip()
            if not message:
                messagebox.showwarning("Cảnh báo", "Vui lòng nhập nội dung cần ký.")
                return

            self.p = int(self.p_var.get())
            self.q = int(self.q_var.get())
            self.h = int(self.h_var.get())
            self.g = int(self.g_var.get())
            self.x = int(self.x_var.get())
            self.k = int(self.k_var.get())
            self.y = int(self.y_var.get())

            if not is_prime(self.p) or not is_prime(self.q):
                raise ValueError("p và q phải là số nguyên tố!")

            # Lấy thuật toán băm được chọn
            hash_alg = self.hash_alg_var.get()

            # Ký với thuật toán băm
            self.r, self.s = sign(message, self.p, self.q, self.g, self.x, self.k, hashAlg=hash_alg)
            self.message = message

            self.signature_text.config(state='normal')
            self.signature_text.delete("1.0", tk.END)
            self.signature_text.insert(tk.END, f"r = {self.r}\ns = {self.s}")
            self.signature_text.config(state='disabled')

        except Exception as e:
            messagebox.showerror("Lỗi khi ký", str(e))

    def save_signature(self):
        if self.r is None or self.s is None:
            messagebox.showwarning("Cảnh báo", "Vui lòng tạo chữ ký trước khi lưu!")
            return
        file_path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text files", "*.txt")])
        if file_path:
            with open(file_path, 'w') as f:
                f.write(f"r = {self.r}\ns = {self.s}")
            messagebox.showinfo("Thành công", f"Chữ ký đã được lưu vào {file_path}")

    def import_signature(self):
        file_path = filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])
        if file_path:
            try:
                with open(file_path, 'r') as f:
                    lines = f.readlines()
                    r = int(lines[0].split('=')[1].strip())
                    s = int(lines[1].split('=')[1].strip())
                self.r_entry.delete(0, tk.END)
                self.r_entry.insert(0, str(r))
                self.s_entry.delete(0, tk.END)
                self.s_entry.insert(0, str(s))
                messagebox.showinfo("Thành công", "Chữ ký đã được nhập thành công!")
            except Exception as e:
                messagebox.showerror("Lỗi", "Không thể nhập chữ ký. Vui lòng kiểm tra file!")

    def verify_signature(self):
        try:
            message = self.verify_text_input.get("1.0", tk.END).strip()
            r2 = int(self.r_entry.get().strip())
            s2 = int(self.s_entry.get().strip())

            self.p = int(self.p_var.get())
            self.q = int(self.q_var.get())
            self.h = int(self.h_var.get())
            self.g = int(self.g_var.get())
            self.y = int(self.y_var.get())

            # Lấy thuật toán băm được chọn
            hash_alg = self.hash_alg_var.get()

            errors = []

            if self.r != r2:
                errors.append("⚠️ r không khớp!")
                self.result_label.config(text="❌ Chữ ký bị thay đổi", fg="red")
            if self.s != s2:
                errors.append("⚠️ s không khớp!")
                self.result_label.config(text="❌ Chữ ký bị thay đổi", fg="red")
            if self.message != message:
                errors.append("⚠️ Văn bản đã bị thay đổi!")
                self.result_label.config(text="❌ Văn bản đã bị thay đổi", fg="red")

            # Xác thực chữ ký với thuật toán băm đã chọn
            is_valid = verify(message, self.p, self.q, self.g, self.y, r2, s2, hashAlg=hash_alg)

            if is_valid:
                self.result_label.config(text="✅ Chữ ký khớp, văn bản toàn vẹn", fg="green")

            if errors:
                messagebox.showwarning("Cảnh báo", "\n".join(errors))

        except Exception as e:
            messagebox.showerror("Lỗi khi xác thực", str(e))

if __name__ == "__main__":
    root = tk.Tk()
    app = DSAGui(root)
    root.mainloop()